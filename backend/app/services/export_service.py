import csv
import io
from datetime import datetime

from fastapi.responses import StreamingResponse
from openpyxl import Workbook


class ExportService:
    async def export_csv(
        self, data: list[dict], columns: list[str], filename: str = "export.csv"
    ) -> StreamingResponse:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in data:
            writer.writerow({k: row.get(k, "") for k in columns})

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    async def export_excel(
        self,
        data: list[dict],
        columns: list[str],
        sheet_name: str = "Sheet1",
        filename: str = "export.xlsx",
        summary_rows: list[dict] | None = None,
        summary_columns: list[str] | None = None,
        building_rows: list[dict] | None = None,
        building_columns: list[str] | None = None,
    ) -> StreamingResponse:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                if isinstance(value, datetime):
                    value = value.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=value)

        if summary_rows and summary_columns:
            ws_summary = wb.create_sheet(title="统计汇总")
            for col_idx, col_name in enumerate(summary_columns, 1):
                ws_summary.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(summary_rows, 2):
                for col_idx, col_name in enumerate(summary_columns, 1):
                    ws_summary.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

        if building_rows and building_columns:
            ws_building = wb.create_sheet(title="分建筑统计")
            for col_idx, col_name in enumerate(building_columns, 1):
                ws_building.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(building_rows, 2):
                for col_idx, col_name in enumerate(building_columns, 1):
                    ws_building.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
