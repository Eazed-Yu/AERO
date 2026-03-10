"""
MCP tool definitions for AERO HVAC full-chain analytics.
Tools cover: catalog, energy queries, chiller/AHU data, COP analysis,
plant efficiency, anomaly lifecycle, equipment inventory, weather, and export.
"""

from datetime import datetime
from typing import Any, Literal
import base64
import csv
import io

from app.mcp.server import MCP_AVAILABLE, is_mcp_enabled

if MCP_AVAILABLE:
    from sqlalchemy import func, select

    from app.database import async_session
    from app.models.energy_meter import EnergyMeter
    from app.models.chiller import ChillerRecord
    from app.models.ahu import AHURecord
    from app.models.weather import WeatherRecord
    from app.mcp.server import mcp
    from app.services.anomaly_service import AnomalyService
    from app.services.building_service import BuildingService
    from app.services.equipment_service import EquipmentService
    from app.services.hvac_data_service import HVACDataService
    from app.services.statistics_service import StatisticsService
    from app.schemas.query import HVACDataQueryParams
    from openpyxl import Workbook

    ALLOWED_PERIODS = {"hour", "day", "week", "month"}
    ENERGY_FIELDS = {
        "id", "building_id", "timestamp",
        "total_electricity_kwh", "hvac_electricity_kwh", "lighting_kwh",
        "plug_load_kwh", "peak_demand_kw", "gas_m3", "water_m3",
        "cooling_kwh", "heating_kwh",
    }
    DEVICE_TYPES = {"chiller", "ahu", "boiler", "vav", "pump", "cooling_tower"}

    def _runtime_disabled_response() -> dict:
        return {"error": "MCP server is disabled at runtime."}

    def _parse_datetime(value: str) -> datetime:
        return datetime.fromisoformat(value)

    def _parse_timerange(start_time: str, end_time: str) -> tuple[datetime, datetime] | tuple[None, None]:
        try:
            start_dt = _parse_datetime(start_time)
            end_dt = _parse_datetime(end_time)
            if start_dt > end_dt:
                return None, None
            return start_dt, end_dt
        except ValueError:
            return None, None

    @mcp.tool()
    async def health_check() -> dict:
        """Return MCP health for connectivity checks."""
        return {
            "ok": is_mcp_enabled(),
            "service": "aero-hvac-mcp",
            "version": "0.2.0",
            "timestamp": datetime.utcnow().isoformat() + "Z",
        }

    @mcp.tool()
    async def get_data_catalog() -> dict:
        """Return field catalog, device types, and query constraints."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        async with async_session() as session:
            buildings = await BuildingService(session).list_buildings()

        return {
            "energy_meter_fields": sorted(ENERGY_FIELDS),
            "device_types": sorted(DEVICE_TYPES),
            "hvac_record_tables": {
                "chiller": ["chw_supply_temp", "chw_return_temp", "chw_flow_rate", "cw_supply_temp", "cw_return_temp", "cw_flow_rate", "power_kw", "cooling_capacity_kw", "load_ratio", "cop"],
                "ahu": ["supply_air_temp", "return_air_temp", "mixed_air_temp", "outdoor_air_temp", "supply_fan_speed", "chw_valve_pos", "hw_valve_pos", "oa_damper_pos", "operating_mode"],
                "boiler": ["hw_supply_temp", "hw_return_temp", "firing_rate", "efficiency", "heating_capacity_kw"],
                "vav": ["zone_temp", "zone_temp_setpoint_clg", "damper_pos", "discharge_air_temp", "zone_co2"],
                "pump": ["speed", "power_kw", "flow_rate", "differential_pressure"],
                "cooling_tower": ["fan_speed", "cw_inlet_temp", "cw_outlet_temp", "wet_bulb_temp", "approach", "range"],
            },
            "periods": sorted(ALLOWED_PERIODS),
            "buildings": [
                {"building_id": b.building_id, "name": b.name, "type": b.building_type, "area": b.area, "cooling_area": b.cooling_area}
                for b in buildings
            ],
        }

    @mcp.tool()
    async def query_energy_data(
        building_id: str | None = None,
        start_time: str | None = None,
        end_time: str | None = None,
        page: int = 1,
        page_size: int = 100,
    ) -> dict:
        """Query building energy meter data with pagination."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        try:
            start_dt = _parse_datetime(start_time) if start_time else None
            end_dt = _parse_datetime(end_time) if end_time else None
        except ValueError as exc:
            return {"error": f"Invalid datetime: {exc}"}

        async with async_session() as session:
            stmt = select(EnergyMeter)
            count_stmt = select(func.count(EnergyMeter.id))

            if building_id:
                stmt = stmt.where(EnergyMeter.building_id == building_id)
                count_stmt = count_stmt.where(EnergyMeter.building_id == building_id)
            if start_dt:
                stmt = stmt.where(EnergyMeter.timestamp >= start_dt)
                count_stmt = count_stmt.where(EnergyMeter.timestamp >= start_dt)
            if end_dt:
                stmt = stmt.where(EnergyMeter.timestamp <= end_dt)
                count_stmt = count_stmt.where(EnergyMeter.timestamp <= end_dt)

            total = (await session.execute(count_stmt)).scalar() or 0
            stmt = stmt.order_by(EnergyMeter.timestamp.desc()).offset((page - 1) * page_size).limit(page_size)
            records = list((await session.execute(stmt)).scalars().all())

        def _row(r):
            return {
                "building_id": r.building_id, "timestamp": r.timestamp.isoformat(),
                "total_electricity_kwh": r.total_electricity_kwh, "hvac_electricity_kwh": r.hvac_electricity_kwh,
                "lighting_kwh": r.lighting_kwh, "plug_load_kwh": r.plug_load_kwh,
                "peak_demand_kw": r.peak_demand_kw, "gas_m3": r.gas_m3, "water_m3": r.water_m3,
                "cooling_kwh": r.cooling_kwh, "heating_kwh": r.heating_kwh,
            }

        return {
            "pagination": {"page": page, "page_size": page_size, "total": total},
            "items": [_row(r) for r in records],
        }

    @mcp.tool()
    async def query_chiller_data(
        device_id: str | None = None,
        start_time: str | None = None,
        end_time: str | None = None,
        page: int = 1,
        page_size: int = 100,
    ) -> dict:
        """Query chiller operating records: COP, temperatures, flow rates, power."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        async with async_session() as session:
            svc = HVACDataService(session)
            params = HVACDataQueryParams(
                device_id=device_id,
                start_time=_parse_datetime(start_time) if start_time else None,
                end_time=_parse_datetime(end_time) if end_time else None,
                page=page, page_size=page_size,
            )
            result = await svc.query_records("chiller", params)

        def _row(r):
            return {
                "device_id": r.device_id, "timestamp": r.timestamp.isoformat(),
                "chw_supply_temp": r.chw_supply_temp, "chw_return_temp": r.chw_return_temp,
                "chw_flow_rate": r.chw_flow_rate, "cw_supply_temp": r.cw_supply_temp,
                "cw_return_temp": r.cw_return_temp, "power_kw": r.power_kw,
                "cooling_capacity_kw": r.cooling_capacity_kw, "cop": r.cop,
                "load_ratio": r.load_ratio, "running_status": r.running_status,
            }

        return {
            "pagination": {"page": result.page, "total": result.total},
            "items": [_row(r) for r in result.items],
        }

    @mcp.tool()
    async def query_ahu_data(
        device_id: str | None = None,
        start_time: str | None = None,
        end_time: str | None = None,
        page: int = 1,
        page_size: int = 100,
    ) -> dict:
        """Query AHU operating records: SAT/RAT/MAT/OAT, valve positions, fan speed."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        async with async_session() as session:
            svc = HVACDataService(session)
            params = HVACDataQueryParams(
                device_id=device_id,
                start_time=_parse_datetime(start_time) if start_time else None,
                end_time=_parse_datetime(end_time) if end_time else None,
                page=page, page_size=page_size,
            )
            result = await svc.query_records("ahu", params)

        def _row(r):
            return {
                "device_id": r.device_id, "timestamp": r.timestamp.isoformat(),
                "supply_air_temp": r.supply_air_temp, "return_air_temp": r.return_air_temp,
                "mixed_air_temp": r.mixed_air_temp, "outdoor_air_temp": r.outdoor_air_temp,
                "supply_fan_speed": r.supply_fan_speed, "chw_valve_pos": r.chw_valve_pos,
                "hw_valve_pos": r.hw_valve_pos, "oa_damper_pos": r.oa_damper_pos,
                "operating_mode": r.operating_mode, "running_status": r.running_status,
            }

        return {
            "pagination": {"page": result.page, "total": result.total},
            "items": [_row(r) for r in result.items],
        }

    @mcp.tool()
    async def analyze_chiller_cop(
        start_time: str,
        end_time: str,
        device_id: str | None = None,
        period: Literal["hour", "day", "week", "month"] = "day",
    ) -> dict:
        """Analyze chiller COP trends based on chiller_records with precise calculations."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        start_dt, end_dt = _parse_timerange(start_time, end_time)
        if not start_dt or not end_dt:
            return {"error": "Invalid datetime range."}

        async with async_session() as session:
            svc = StatisticsService(session)
            results = await svc.calculate_cop(start_dt, end_dt, device_id=device_id, period=period)

        return {
            "period": period, "count": len(results),
            "items": [
                {
                    "period_start": r.period_start.isoformat(), "device_id": r.device_id,
                    "cop": r.cop, "cooling_capacity_kwh": r.cooling_capacity_kwh,
                    "power_kwh": r.power_kwh, "chw_supply_temp_avg": r.chw_supply_temp_avg,
                    "chw_return_temp_avg": r.chw_return_temp_avg, "load_ratio_avg": r.load_ratio_avg,
                    "rating": r.rating,
                }
                for r in results
            ],
        }

    @mcp.tool()
    async def analyze_plant_efficiency(
        start_time: str,
        end_time: str,
        period: Literal["hour", "day", "week", "month"] = "day",
    ) -> dict:
        """Analyze cooling plant system efficiency (chiller + pump + tower combined COP)."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        start_dt, end_dt = _parse_timerange(start_time, end_time)
        if not start_dt or not end_dt:
            return {"error": "Invalid datetime range."}

        async with async_session() as session:
            svc = StatisticsService(session)
            results = await svc.plant_efficiency(start_dt, end_dt, period)

        return {
            "period": period, "count": len(results),
            "items": [
                {
                    "period_start": r.period_start.isoformat(),
                    "total_cooling_kwh": r.total_cooling_kwh,
                    "chiller_power_kwh": r.chiller_power_kwh,
                    "pump_power_kwh": r.pump_power_kwh,
                    "tower_power_kwh": r.tower_power_kwh,
                    "system_cop": r.system_cop,
                }
                for r in results
            ],
        }

    @mcp.tool()
    async def get_weather_data(
        building_id: str,
        start_time: str,
        end_time: str,
        page: int = 1,
        page_size: int = 100,
    ) -> dict:
        """Query weather records for a building location."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        start_dt, end_dt = _parse_timerange(start_time, end_time)
        if not start_dt or not end_dt:
            return {"error": "Invalid datetime range."}

        async with async_session() as session:
            stmt = select(WeatherRecord).where(
                WeatherRecord.building_id == building_id,
                WeatherRecord.timestamp >= start_dt,
                WeatherRecord.timestamp <= end_dt,
            ).order_by(WeatherRecord.timestamp.desc()).offset((page - 1) * page_size).limit(page_size)
            count_stmt = select(func.count(WeatherRecord.id)).where(
                WeatherRecord.building_id == building_id,
                WeatherRecord.timestamp >= start_dt,
                WeatherRecord.timestamp <= end_dt,
            )
            total = (await session.execute(count_stmt)).scalar() or 0
            records = list((await session.execute(stmt)).scalars().all())

        return {
            "building_id": building_id,
            "pagination": {"page": page, "total": total},
            "items": [
                {
                    "timestamp": r.timestamp.isoformat(),
                    "dry_bulb_temp": r.dry_bulb_temp, "wet_bulb_temp": r.wet_bulb_temp,
                    "relative_humidity": r.relative_humidity, "wind_speed": r.wind_speed,
                    "solar_radiation": r.solar_radiation,
                }
                for r in records
            ],
        }

    @mcp.tool()
    async def run_anomaly_detection(
        building_id: str, start_time: str, end_time: str,
    ) -> dict:
        """Run anomaly detection across energy and HVAC data, persist results."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        start_dt, end_dt = _parse_timerange(start_time, end_time)
        if not start_dt or not end_dt:
            return {"error": "Invalid datetime range."}

        async with async_session() as session:
            svc = AnomalyService(session)
            events = await svc.detect_anomalies(building_id, start_dt, end_dt)
            await session.commit()

        return {
            "detected_count": len(events),
            "items": [
                {
                    "id": e.id, "timestamp": e.timestamp.isoformat(),
                    "anomaly_type": e.anomaly_type, "severity": e.severity,
                    "metric_name": e.metric_name, "metric_value": e.metric_value,
                    "description": e.description, "equipment_type": e.equipment_type,
                    "fault_code": e.fault_code, "recommended_action": e.recommended_action,
                }
                for e in events
            ],
        }

    @mcp.tool()
    async def query_anomaly_events(
        building_id: str | None = None,
        severity: Literal["low", "medium", "high", "critical"] | None = None,
        equipment_type: str | None = None,
        resolved: bool | None = None,
        start_time: str | None = None,
        end_time: str | None = None,
        limit: int = 200,
    ) -> dict:
        """Query existing anomaly events with filters."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        try:
            start_dt = _parse_datetime(start_time) if start_time else None
            end_dt = _parse_datetime(end_time) if end_time else None
        except ValueError as exc:
            return {"error": f"Invalid datetime: {exc}"}

        async with async_session() as session:
            svc = AnomalyService(session)
            items = await svc.list_anomalies(
                building_id=building_id, severity=severity, resolved=resolved,
                equipment_type=equipment_type, start_time=start_dt, end_time=end_dt, limit=limit,
            )

        return {
            "count": len(items),
            "items": [
                {
                    "id": e.id, "building_id": e.building_id, "device_id": e.device_id,
                    "timestamp": e.timestamp.isoformat(), "anomaly_type": e.anomaly_type,
                    "severity": e.severity, "description": e.description,
                    "equipment_type": e.equipment_type, "fault_code": e.fault_code,
                    "resolved": e.resolved,
                }
                for e in items
            ],
        }

    @mcp.tool()
    async def query_equipment_inventory(
        building_id: str | None = None,
        device_type: str | None = None,
        system_type: str | None = None,
    ) -> dict:
        """Query equipment inventory (device ledger)."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        async with async_session() as session:
            svc = EquipmentService(session)
            equipments = await svc.list_equipment(building_id=building_id, device_type=device_type, system_type=system_type)

        return {
            "count": len(equipments),
            "items": [
                {
                    "device_id": eq.device_id, "device_name": eq.device_name,
                    "device_type": eq.device_type, "system_type": eq.system_type,
                    "building_id": eq.building_id, "model": eq.model,
                    "manufacturer": eq.manufacturer, "rated_power_kw": eq.rated_power_kw,
                    "rated_capacity": eq.rated_capacity, "rated_cop": eq.rated_cop,
                    "status": eq.status,
                }
                for eq in equipments
            ],
        }

    @mcp.tool()
    async def export_energy_dataset(
        building_id: str, start_time: str, end_time: str,
        format: Literal["csv", "excel"] = "csv",
    ) -> dict:
        """Export energy dataset as base64 payload."""
        if not is_mcp_enabled():
            return _runtime_disabled_response()

        start_dt, end_dt = _parse_timerange(start_time, end_time)
        if not start_dt or not end_dt:
            return {"error": "Invalid datetime range."}

        export_fields = [
            "building_id", "timestamp", "total_electricity_kwh", "hvac_electricity_kwh",
            "lighting_kwh", "plug_load_kwh", "gas_m3", "water_m3", "cooling_kwh", "heating_kwh",
        ]

        async with async_session() as session:
            stmt = (
                select(EnergyMeter)
                .where(EnergyMeter.building_id == building_id)
                .where(EnergyMeter.timestamp >= start_dt)
                .where(EnergyMeter.timestamp <= end_dt)
                .order_by(EnergyMeter.timestamp.asc())
            )
            records = list((await session.execute(stmt)).scalars().all())

        rows = []
        for r in records:
            row = {}
            for f in export_fields:
                v = getattr(r, f, None)
                row[f] = v.isoformat() if isinstance(v, datetime) else v
            rows.append(row)

        if format == "csv":
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=export_fields)
            writer.writeheader()
            writer.writerows(rows)
            content_bytes = output.getvalue().encode("utf-8")
            filename = f"energy_{building_id}.csv"
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "energy"
            for ci, cn in enumerate(export_fields, 1):
                ws.cell(row=1, column=ci, value=cn)
            for ri, rd in enumerate(rows, 2):
                for ci, cn in enumerate(export_fields, 1):
                    ws.cell(row=ri, column=ci, value=rd.get(cn))
            stream = io.BytesIO()
            wb.save(stream)
            content_bytes = stream.getvalue()
            filename = f"energy_{building_id}.xlsx"

        return {
            "filename": filename,
            "record_count": len(rows),
            "content_base64": base64.b64encode(content_bytes).decode("ascii"),
        }
